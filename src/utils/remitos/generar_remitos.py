import datetime
import os
from tkinter import filedialog, messagebox
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, Border, Side
from database import session, Clientes

def generar_remito_excel(cliente, carrito, observaciones, imprimir):
    # Verificar si se seleccionó un cliente
    if not cliente:
        return

    # Obtener la ruta del escritorio del usuario actual
    desktop_path = os.path.join(os.path.expanduser("~"), "Desktop")

    # Construir la ruta relativa a la carpeta "NOTA DE ENTREGA" en el escritorio
    nota_de_entrega_path = os.path.join(desktop_path, "NOTA DE ENTREGA")

    # Solicitar al usuario la ubicación donde guardar el remito, con la carpeta por defecto NOTA DE ENTREGA
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        initialdir=nota_de_entrega_path,
        title="Guardar remito como..."
    )

    if not file_path:
        return
    
    # Cargar la plantilla de Excel para el remito con openpyxl y obtener la hoja activa
    wb = load_workbook('./data/PLANTILLA REMITO.xlsx')
    sheet = wb.active

    # Rellenar los datos del remito

    # Obtener la fecha actual
    fecha_actual = datetime.date.today().strftime("%d-%m-%Y")

    # Si el cliente es "Consumidor Final", no imputamos datos de cliente
    if cliente != "Consumidor Final":
        # Verificar si el cliente está en la base de datos
        cliente_db = session.query(Clientes).filter_by(nombre=cliente).first()
        if cliente_db:
            nombre_cliente = cliente_db.nombre
            cuit_cliente = cliente_db.cuit
            domicilio_cliente = cliente_db.direccion
            telefono_cliente = cliente_db.telefono

            # Escribir los datos del cliente en las celdas correspondientes
            sheet.cell(row=5, column=4, value=f"CLIENTE: {nombre_cliente}").font = Font(name='Arial', size=12)
            sheet.cell(row=5, column=6, value=f"DOMICILIO: {domicilio_cliente}").font = Font(name='Arial', size=12)
            sheet.cell(row=7, column=6, value=f"CUIT: {cuit_cliente}").font = Font(name='Arial', size=12)
            sheet.cell(row=7, column=4, value=f"TELÉFONO: {telefono_cliente}").font = Font(name='Arial', size=12)
            sheet.cell(row=3, column=4, value=f"Fecha de entrega: {fecha_actual}").font = Font(name='Arial', size=12)

    
    else:
        sheet.cell(row=4, column=4, value="CONSUMIDOR FINAL").font = Font(name='Arial', size=12)
        sheet.cell(row=4, column=7, value=f"Fecha de entrega: {fecha_actual}").font = Font(name='Arial', size=12)


    fila_inicial = 10
    total = 0

    # Iterar sobre los productos en el carrito y hacer los cálculos necesarios
    for item in carrito:
        producto, cantidad, descuento, precio = item
        # Convertir la cantidad y el precio a entero y flotante
        cantidad = float(cantidad)
        # Quitamos el signo de pesos y las comas
        precio = precio[1:].replace(',', '') if precio else 0
        precio = float(precio) if precio else 0
        # Quitar el signo de porcentaje y convertir a flotante el descuento
        descuento = float(descuento[:-1]) if descuento else ''
        # Calcular el precio con descuento
        precio = precio * (1 - descuento / 100 if descuento else 1)
        # Calcular el precio total del producto
        precio_total = cantidad * precio

        # Escribir los datos en las celdas correspondientes
        sheet.cell(row=fila_inicial, column=1, value=cantidad)
        sheet.cell(row=fila_inicial, column=2, value=producto)
        sheet.cell(row=fila_inicial, column=7, value=precio if precio > 0 else "")
        sheet.cell(row=fila_inicial, column=8, value=precio_total if precio_total > 0 else "")

        # Pasar a la siguiente fila
        total += precio_total
        fila_inicial += 1
    
    # Imputar el total
    sheet.cell(row=fila_inicial +1, column=8, value=total if total > 0 else "").font = Font(name='Arial', size=12, bold=True)

    # Firma del cliente y observaciones
    sheet.cell(row=fila_inicial + 3, column=1, value="FIRMA DEL CLIENTE:").font = Font(name='Arial', size=12, bold=True)
    sheet.cell(row=fila_inicial + 3, column=5, value=f"OBSERVACIONES: {observaciones}").font = Font(name='Arial', size=12, bold=True)

    # Alinear firma del cliente a la izquierda
    sheet.cell(row=fila_inicial + 3, column=1).alignment = Alignment(horizontal='left')

    # Copia para la empresa

    if fila_inicial <= 40:
        fila_inicial = 44
        # Fusionar celdas para el título
        sheet.merge_cells(start_row=fila_inicial, start_column=7, end_row=fila_inicial +2, end_column=8)
        # Titulo "REMITO" con fuente 'Arial' de tamaño 20 y negrita
        sheet.cell(row=fila_inicial, column=7, value="REMITO").font = Font(name='Arial', size=26, bold=True, color="4f81bd")
        # Alinear a la derecha
        sheet.cell(row=fila_inicial, column=7).alignment = Alignment(horizontal='right', vertical='top')

        # Etiqueta "Fecha de entrega"
        sheet.cell(row=fila_inicial + 1, column=5, value=f"Fecha de entrega: {fecha_actual}")

        # Si el cliente no es consumidor final, rellenar
        if cliente != "Consumidor Final":
            # Rellenar con los datos del cliente
            sheet.cell(row=fila_inicial + 1, column=1, value=f"CLIENTE: {nombre_cliente}").font = Font(name='Arial', size=12)
            sheet.cell(row=fila_inicial + 1, column=3, value=f"CUIT: {cuit_cliente}").font = Font(name='Arial', size=12)
            sheet.cell(row=fila_inicial + 3, column=3, value=f"DOMICILIO: {domicilio_cliente}").font = Font(name='Arial', size=12)
            sheet.cell(row=fila_inicial + 3, column=1, value=f"TELÉFONO: {telefono_cliente}").font = Font(name='Arial', size=12)
            # Alinear a la izquierda Cliente y Domicilio
            sheet.cell(row=fila_inicial + 1, column=1).alignment = Alignment(horizontal='left')
            sheet.cell(row=fila_inicial + 3, column=1).alignment = Alignment(horizontal='left')

        else:
            sheet.cell(row=fila_inicial + 1, column=1, value="CLIENTE: CONSUMIDOR FINAL").font = Font(name='Arial', size=12)
            # Alinear a la izquierda 
            sheet.cell(row=fila_inicial + 1, column=1).alignment = Alignment(horizontal='left')
            fila_inicial -= 2

        
        # Alinear a la izquierda Cliente y Domicilio
        sheet.cell(row=fila_inicial + 1, column=1).alignment = Alignment(horizontal='left')
        sheet.cell(row=fila_inicial + 3, column=1).alignment = Alignment(horizontal='left')

        # Borde para las celdas
        borde_cant = Border(left=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
        
        borde_topbot = Border(top=Side(style='thin'),
                            bottom=Side(style='thin'))
        
        borde_total = Border(top=Side(style='thin'),
                            bottom=Side(style='thin'),
                            right=Side(style='thin'))

        # CANTIDAD, DETALLE, PRECIO UD., TOTAL
        sheet.cell(row=fila_inicial + 5, column=1, value="CANTIDAD").border = borde_cant
        sheet.cell(row=fila_inicial + 5, column=2, value="DETALLE").border = borde_topbot
        for i in range(3, 7):
            sheet.cell(row=fila_inicial + 5, column=i).border = borde_topbot
        sheet.cell(row=fila_inicial + 5, column=7, value="PRECIO UD.").border = borde_topbot
        sheet.cell(row=fila_inicial + 5, column=8, value="TOTAL").border = borde_total

        # Aplicar negrita a la fila de los encabezados
        sheet.cell(row=fila_inicial + 5, column=1).font = Font(name='Arial', size=12, bold=True)
        sheet.cell(row=fila_inicial + 5, column=2).font = Font(name='Arial', size=12, bold=True)
        sheet.cell(row=fila_inicial + 5, column=7).font = Font(name='Arial', size=12, bold=True)
        sheet.cell(row=fila_inicial + 5, column=8).font = Font(name='Arial', size=12, bold=True)

        fila_inicial += 6

        # Imputar los datos del carrito
        for item in carrito:
            producto, cantidad, descuento, precio = item
            # Convertir la cantidad y el precio a flotante
            cantidad = float(cantidad)
            # Quitamos el signo de pesos y las comas
            precio = precio[1:].replace(',', '') if precio else 0
            precio = float(precio) if precio else 0
            # Quitar el signo de porcentaje y convertir a flotante el descuento
            descuento = float(descuento[:-1]) if descuento else ''
            # Calcular el precio con descuento
            precio = precio * (1 - descuento / 100 if descuento else 1)
            # Calcular el precio total del producto
            precio_total = cantidad * precio

            # Escribir los datos en las celdas correspondientes

            sheet.cell(row=fila_inicial, column=1, value=cantidad)
            sheet.cell(row=fila_inicial, column=2, value=producto)
            sheet.cell(row=fila_inicial, column=7, value=precio if precio > 0 else "")
            sheet.cell(row=fila_inicial, column=8, value=precio_total if precio_total > 0 else "")

            # Pasar a la siguiente fila
            fila_inicial += 1
        
        # Imputar el total
        sheet.cell(row=fila_inicial + 1, column=8, value=total if total > 0 else "").font = Font(name='Arial', size=12, bold=True)

        # Firma del cliente y observaciones
        sheet.cell(row=fila_inicial + 3, column=1, value="FIRMA DEL CLIENTE:").font = Font(name='Arial', size=12, bold=True)
        sheet.cell(row=fila_inicial + 3, column=5, value=f"OBSERVACIONES: {observaciones}").font = Font(name='Arial', size=12, bold=True)

        # Alinear firma del cliente a la izquierda
        sheet.cell(row=fila_inicial + 3, column=1).alignment = Alignment(horizontal='left')

    # Guardar el archivo Excel
    wb.save(file_path)
    messagebox.showinfo("Éxito", f"Remito generado en {file_path}")

    # Si clickeó en imprimir, imprimir el remito
    if imprimir:
        try:
            os.startfile(file_path, "print")
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo imprimir el remito: {str(e)}")

    else:
        # Abrir el archivo automáticamente
        try:
            os.startfile(file_path)

        except Exception as e:
            messagebox.showerror("Error", f"No se pudo abrir el archivo: {str(e)}")